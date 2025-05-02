import os
import logging
from typing import Dict, List, Any, Optional, Tuple, Union
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
import xlrd
import csv

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelProcessor:
    """
    Excel/CSV Processor for extracting data and structure from financial spreadsheets.
    """
    
    def __init__(self):
        """
        Initialize the Excel Processor.
        """
        pass
    
    def process_file(self, file_path: str) -> Dict[str, Any]:
        """
        Process an Excel or CSV file and extract its content.
        
        Args:
            file_path: Path to the Excel or CSV file
            
        Returns:
            Dictionary containing extracted content
        """
        logger.info(f"Processing spreadsheet: {file_path}")
        
        try:
            # Check if file exists
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            # Determine file type
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext in ['.xlsx', '.xlsm', '.xls']:
                return self._process_excel(file_path, file_ext)
            elif file_ext == '.csv':
                return self._process_csv(file_path)
            else:
                raise ValueError(f"Unsupported file type: {file_ext}")
        except Exception as e:
            logger.error(f"Error processing file {file_path}: {e}")
            raise
    
    def _process_excel(self, file_path: str, file_ext: str) -> Dict[str, Any]:
        """
        Process an Excel file.
        
        Args:
            file_path: Path to the Excel file
            file_ext: File extension
            
        Returns:
            Dictionary containing extracted content
        """
        # Use different libraries based on file extension
        if file_ext in ['.xlsx', '.xlsm']:
            return self._process_xlsx(file_path)
        else:  # .xls
            return self._process_xls(file_path)
    
    def _process_xlsx(self, file_path: str) -> Dict[str, Any]:
        """
        Process an XLSX file using openpyxl.
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            Dictionary containing extracted content
        """
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Extract basic metadata
            metadata = {
                "file_path": file_path,
                "file_name": os.path.basename(file_path),
                "sheet_names": workbook.sheetnames,
                "properties": self._extract_xlsx_properties(workbook)
            }
            
            # Process each sheet
            sheets = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = self._process_xlsx_sheet(sheet)
                sheets.append(sheet_data)
            
            # Extract named ranges
            named_ranges = self._extract_xlsx_named_ranges(workbook)
            
            # Extract formulas
            formulas = self._extract_xlsx_formulas(workbook)
            
            return {
                "metadata": metadata,
                "sheets": sheets,
                "named_ranges": named_ranges,
                "formulas": formulas
            }
        except Exception as e:
            logger.error(f"Error processing XLSX file: {e}")
            raise
    
    def _process_xlsx_sheet(self, sheet) -> Dict[str, Any]:
        """
        Process a single sheet from an XLSX file.
        
        Args:
            sheet: openpyxl worksheet
            
        Returns:
            Dictionary containing sheet data
        """
        # Get sheet dimensions
        min_row, min_col, max_row, max_col = self._get_sheet_dimensions(sheet)
        
        # Extract data
        data = []
        for row in range(min_row, max_row + 1):
            row_data = []
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                row_data.append(self._get_cell_value(cell))
            data.append(row_data)
        
        # Convert to DataFrame for easier processing
        df = pd.DataFrame(data)
        
        # Detect header row
        header_row = self._detect_header_row(df)
        
        # Use header row if found
        if header_row is not None:
            headers = df.iloc[header_row].tolist()
            df = df.iloc[header_row + 1:].reset_index(drop=True)
            df.columns = headers
        
        # Clean data
        df = self._clean_dataframe(df)
        
        # Detect tables within the sheet
        tables = self._detect_tables(df)
        
        # Extract sheet structure
        structure = self._extract_sheet_structure(sheet)
        
        return {
            "name": sheet.title,
            "data": df.values.tolist(),
            "headers": df.columns.tolist(),
            "tables": tables,
            "structure": structure,
            "dimensions": {
                "min_row": min_row,
                "min_col": min_col,
                "max_row": max_row,
                "max_col": max_col
            }
        }
    
    def _get_sheet_dimensions(self, sheet) -> Tuple[int, int, int, int]:
        """
        Get the dimensions of a sheet.
        
        Args:
            sheet: openpyxl worksheet
            
        Returns:
            Tuple of (min_row, min_col, max_row, max_col)
        """
        min_row, min_col = 1, 1
        max_row, max_col = 1, 1
        
        if sheet.max_row > 1:
            max_row = sheet.max_row
        
        if sheet.max_column > 1:
            max_col = sheet.max_column
        
        # Find actual data boundaries
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None:
                    min_row = min(min_row, row)
                    min_col = min(min_col, col)
                    max_row = max(max_row, row)
                    max_col = max(max_col, col)
        
        return min_row, min_col, max_row, max_col
    
    def _get_cell_value(self, cell) -> Any:
        """
        Get the value of a cell.
        
        Args:
            cell: openpyxl cell
            
        Returns:
            Cell value
        """
        if cell.value is None:
            return ""
        
        # Return the cell value
        return cell.value
    
    def _detect_header_row(self, df: pd.DataFrame) -> Optional[int]:
        """
        Detect the header row in a DataFrame.
        
        Args:
            df: DataFrame
            
        Returns:
            Index of header row or None if not found
        """
        # Check first 10 rows
        for i in range(min(10, len(df))):
            row = df.iloc[i]
            
            # Check if row has mostly strings
            string_ratio = sum(isinstance(val, str) for val in row) / max(len(row), 1)
            
            # Check if strings are short (typical for headers)
            avg_len = sum(len(str(val)) for val in row) / max(len(row), 1)
            
            # Check if row has few empty cells
            empty_ratio = sum(pd.isna(val) or val == "" for val in row) / max(len(row), 1)
            
            if string_ratio > 0.7 and avg_len < 30 and empty_ratio < 0.3:
                return i
        
        return None
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean a DataFrame.
        
        Args:
            df: DataFrame
            
        Returns:
            Cleaned DataFrame
        """
        # Replace NaN with empty string
        df = df.fillna("")
        
        # Convert all values to strings
        df = df.applymap(lambda x: str(x).strip() if isinstance(x, str) else x)
        
        # Remove empty rows
        df = df[df.astype(str).apply(lambda x: x.str.strip().str.len() > 0).any(axis=1)]
        
        # Remove empty columns
        df = df.loc[:, df.astype(str).apply(lambda x: x.str.strip().str.len() > 0).any()]
        
        return df
    
    def _detect_tables(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Detect tables within a DataFrame.
        
        Args:
            df: DataFrame
            
        Returns:
            List of detected tables
        """
        tables = []
        
        # Simple heuristic: Look for blocks of data with consistent structure
        # This is a simplified approach; a more sophisticated algorithm would be needed for complex sheets
        
        # Check if the entire DataFrame looks like a table
        if len(df) > 1 and len(df.columns) > 1:
            # Check if first row looks like headers
            first_row = df.iloc[0]
            if all(isinstance(val, str) for val in first_row):
                tables.append({
                    "name": "Table 1",
                    "start_row": 0,
                    "end_row": len(df) - 1,
                    "start_col": 0,
                    "end_col": len(df.columns) - 1,
                    "headers": df.columns.tolist()
                })
        
        return tables
    
    def _extract_sheet_structure(self, sheet) -> Dict[str, Any]:
        """
        Extract the structure of a sheet.
        
        Args:
            sheet: openpyxl worksheet
            
        Returns:
            Dictionary containing sheet structure
        """
        # Extract merged cells
        merged_cells = []
        for merged_range in sheet.merged_cells.ranges:
            merged_cells.append({
                "start_row": merged_range.min_row,
                "end_row": merged_range.max_row,
                "start_col": merged_range.min_col,
                "end_col": merged_range.max_col
            })
        
        # Extract column widths
        column_widths = {}
        for i, column_letter in enumerate(get_column_letter(i) for i in range(1, sheet.max_column + 1)):
            if column_letter in sheet.column_dimensions:
                width = sheet.column_dimensions[column_letter].width
                if width is not None:
                    column_widths[i] = width
        
        # Extract row heights
        row_heights = {}
        for i in range(1, sheet.max_row + 1):
            if i in sheet.row_dimensions:
                height = sheet.row_dimensions[i].height
                if height is not None:
                    row_heights[i] = height
        
        return {
            "merged_cells": merged_cells,
            "column_widths": column_widths,
            "row_heights": row_heights
        }
    
    def _extract_xlsx_properties(self, workbook) -> Dict[str, Any]:
        """
        Extract properties from an XLSX workbook.
        
        Args:
            workbook: openpyxl workbook
            
        Returns:
            Dictionary of properties
        """
        properties = {}
        
        if workbook.properties:
            props = workbook.properties
            if props.title:
                properties["title"] = props.title
            if props.subject:
                properties["subject"] = props.subject
            if props.creator:
                properties["creator"] = props.creator
            if props.created:
                properties["created"] = props.created.isoformat()
            if props.modified:
                properties["modified"] = props.modified.isoformat()
        
        return properties
    
    def _extract_xlsx_named_ranges(self, workbook) -> List[Dict[str, Any]]:
        """
        Extract named ranges from an XLSX workbook.
        
        Args:
            workbook: openpyxl workbook
            
        Returns:
            List of named ranges
        """
        named_ranges = []
        
        for name in workbook.defined_names.definedName:
            named_ranges.append({
                "name": name.name,
                "value": name.value
            })
        
        return named_ranges
    
    def _extract_xlsx_formulas(self, workbook) -> List[Dict[str, Any]]:
        """
        Extract formulas from an XLSX workbook.
        
        Args:
            workbook: openpyxl workbook
            
        Returns:
            List of formulas
        """
        formulas = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.formula:
                        formulas.append({
                            "sheet": sheet_name,
                            "row": row,
                            "col": col,
                            "address": f"{get_column_letter(col)}{row}",
                            "formula": cell.formula
                        })
        
        return formulas
    
    def _process_xls(self, file_path: str) -> Dict[str, Any]:
        """
        Process an XLS file using xlrd.
        
        Args:
            file_path: Path to the XLS file
            
        Returns:
            Dictionary containing extracted content
        """
        try:
            # Load workbook
            workbook = xlrd.open_workbook(file_path)
            
            # Extract basic metadata
            metadata = {
                "file_path": file_path,
                "file_name": os.path.basename(file_path),
                "sheet_names": workbook.sheet_names()
            }
            
            # Process each sheet
            sheets = []
            for sheet_name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(sheet_name)
                sheet_data = self._process_xls_sheet(sheet)
                sheets.append(sheet_data)
            
            return {
                "metadata": metadata,
                "sheets": sheets,
                "named_ranges": [],  # xlrd doesn't easily expose named ranges
                "formulas": []  # xlrd doesn't easily expose formulas
            }
        except Exception as e:
            logger.error(f"Error processing XLS file: {e}")
            raise
    
    def _process_xls_sheet(self, sheet) -> Dict[str, Any]:
        """
        Process a single sheet from an XLS file.
        
        Args:
            sheet: xlrd sheet
            
        Returns:
            Dictionary containing sheet data
        """
        # Extract data
        data = []
        for row in range(sheet.nrows):
            row_data = []
            for col in range(sheet.ncols):
                cell_value = sheet.cell_value(row, col)
                cell_type = sheet.cell_type(row, col)
                
                # Handle date values
                if cell_type == xlrd.XL_CELL_DATE:
                    try:
                        cell_value = xlrd.xldate.xldate_as_datetime(cell_value, workbook.datemode)
                    except:
                        pass
                
                row_data.append(cell_value)
            data.append(row_data)
        
        # Convert to DataFrame for easier processing
        df = pd.DataFrame(data)
        
        # Detect header row
        header_row = self._detect_header_row(df)
        
        # Use header row if found
        if header_row is not None:
            headers = df.iloc[header_row].tolist()
            df = df.iloc[header_row + 1:].reset_index(drop=True)
            df.columns = headers
        
        # Clean data
        df = self._clean_dataframe(df)
        
        # Detect tables within the sheet
        tables = self._detect_tables(df)
        
        return {
            "name": sheet.name,
            "data": df.values.tolist(),
            "headers": df.columns.tolist(),
            "tables": tables,
            "dimensions": {
                "rows": sheet.nrows,
                "cols": sheet.ncols
            }
        }
    
    def _process_csv(self, file_path: str) -> Dict[str, Any]:
        """
        Process a CSV file.
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            Dictionary containing extracted content
        """
        try:
            # Detect delimiter
            delimiter = self._detect_csv_delimiter(file_path)
            
            # Read CSV file
            df = pd.read_csv(file_path, delimiter=delimiter, encoding='utf-8', error_bad_lines=False)
            
            # Clean data
            df = self._clean_dataframe(df)
            
            # Extract metadata
            metadata = {
                "file_path": file_path,
                "file_name": os.path.basename(file_path),
                "delimiter": delimiter,
                "num_rows": len(df),
                "num_cols": len(df.columns)
            }
            
            # Detect tables
            tables = self._detect_tables(df)
            
            return {
                "metadata": metadata,
                "sheets": [{
                    "name": "Sheet1",
                    "data": df.values.tolist(),
                    "headers": df.columns.tolist(),
                    "tables": tables,
                    "dimensions": {
                        "rows": len(df),
                        "cols": len(df.columns)
                    }
                }],
                "named_ranges": [],
                "formulas": []
            }
        except Exception as e:
            logger.error(f"Error processing CSV file: {e}")
            raise
    
    def _detect_csv_delimiter(self, file_path: str) -> str:
        """
        Detect the delimiter used in a CSV file.
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            Detected delimiter
        """
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            sample = f.read(4096)
        
        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(sample)
            return dialect.delimiter
        except:
            # Default to comma if detection fails
            return ','
